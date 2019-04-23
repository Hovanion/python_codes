import openpyxl, os
from pprint import pprint

THIS_FOLDER = os.path.dirname(os.path.abspath(__file__))
print(THIS_FOLDER)

THIS_FOLDER+="\\xlsxek"
folder=os.listdir(THIS_FOLDER)
adat=dict()
for file in folder:
    my_file = os.path.join(THIS_FOLDER, file)
    print (my_file)
    book = openpyxl.load_workbook(my_file)
    #print(book.get_sheet_names())
    #sheet = book.get_sheet_by_name("Oldal1")
    #print(book.get_sheet_names())
    #print(sheet.title)

    active_sheet = book.active
    print(type(active_sheet))

    #sheet.sheet_properties.tabColor = "0072BA"

    #book.save('sheets3.xlsx')
    active_sheet = book.active


    b1=active_sheet['B3']
    #print (b1.value)
    b3=active_sheet.cell(row=3,column=2)
    #print (b3.value)

    #print(active_sheet.max_row)

    sheet_obj = book.active
    m_row = sheet_obj.max_row

    # Loop will print all values
    # of first column
    for i in range(1, m_row + 1):
        cell_obj = sheet_obj.cell(row = i, column = 1)
        #print(cell_obj.value)

    '''
    print (active_sheet['B3'].fill.bgColor, active_sheet['B3'].value)
    print ("#", active_sheet['A1'].fill.fgColor,  active_sheet['D3'].fill.bgColor, active_sheet['D3'].value)
    print ("a3", active_sheet['A3'].fill.start_color.index, active_sheet['A3'].fill.bgColor,  active_sheet['A3'].value)
    print ("-----------")
    print (active_sheet['A1'].fill.fgColor, active_sheet['B3'].fill.fgColor, active_sheet['C3'].fill.fgColor, active_sheet['D3'].fill.fgColor.tint)
    '''

    cells=["B3", "C3", "D3", "E3", "F3"]

    for c in cells:
        if active_sheet[c].fill.fgColor.tint != 0:
            #print (c)
            print (active_sheet[c].value)


    labels=dict()

    print ("--------------")
    nev=active_sheet["B1"].value
    adat[nev]=dict()
    #adat[active_sheet["B2"].value]=dict()
    for i in range(1,sheet_obj.max_row, 2):
        #nev=active_sheet.cell(row = i, column = 2).value

        #print(active_sheet.cell(row = i, column = 1).value)

        hivatkozas="A"+str(i)
        ertek=active_sheet.cell(row = i, column = 2).value
        cella=active_sheet.cell(row = i, column = 1).value

        if i==3:
            cells=["B3", "C3", "D3", "E3", "F3"]
            print (sheet_obj.max_column)

            for c in cells:
                if active_sheet[c].fill.fgColor.tint != 0:
                    print (c)
                    print ("Ez van a 3. sorban", active_sheet[c].value)

            for d in range(2,sheet_obj.max_column):
                if active_sheet.cell(row=i, column=d).fill.fgColor.tint != 0:
                    print (active_sheet.cell(row=i, column=d))
                    print ("Ez van a 3. sorban", active_sheet.cell(row=i, column=d).value)
                    #adat[nev]["A"+str(i)]=active_sheet.cell(row=i, column=d).value
                    adat[nev][cella]=active_sheet.cell(row=i, column=d).value
        else:
            if ertek== None:
                ertek="XXXX"
            adat[nev][cella]=ertek





pprint (adat)
